#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpasen import common, api
import random
import os

def init():
    try:
        from openpyxl import load_workbook
        global wb
        wb = load_workbook(filename="assets/templates/horario.xlsx")
        reportError = False
        return reportError
    except ImportError:
        print("No tienes instalado openpyxl")
        reportError = True
        return reportError

# https://blog.stigok.com/2019/06/22/generate-random-color-codes-python.html
def rand_web_color_hex():
    rgb = ""
    for _ in "RGB":
        i = random.randrange(96, 2**8)
        rgb += i.to_bytes(1, "big").hex()
    return rgb

def generar(nombre, curso):
    reportError = init()
    # Parar si el usuario no tiene bs4 instalado
    if reportError is True:
        return "ImportError"
    
    # Imports adicionales
    from openpyxl.styles import PatternFill, Alignment, Border, Side

    ws = wb.active
    # Conseguir horario de la API
    horario = api.horario()
    asignaturas = {}
    for col, dia in zip(ws.iter_cols(min_row=2, max_row=7, min_col=2, max_col=6), horario):
        for cell in col:
            asignatura = horario[dia][col.index(cell)]
            cell.value = asignatura

            # Color celda
            # Si ya hay un color asignado a la asignatura, usarlo
            if asignatura in asignaturas:
                color = asignaturas[asignatura]
            # Si no hay un color asignado elige un color al azar
            else:
                color = rand_web_color_hex()
                asignaturas[asignatura] = color
            
            # Rellenar celda con color
            cell.fill = PatternFill(start_color=color, fill_type = "solid")
    
    # Crear row con separación de recreo

    # Bordes

    thin = Side(border_style="thin", color="000000")
    # Damn boy he
    thick = Side(border_style='thick', color='000000')
    # Boy

    ws.insert_rows(5)
    ws.merge_cells('B5:F5')
    ws['B5'] = "RECREO"
    ws['B5'].border = Border(right=thick, top=thin, bottom=thin)
    ws['B5'].alignment = Alignment(horizontal='center', vertical='center')
    # Curso en la esquina
    ws['A1'] = curso

    # Rellenar color recreo izquierda
    ws['A5'].fill = PatternFill(start_color="FF94BD5E", fill_type="solid")

    try:
        os.mkdir(f'{common.config_path}{nombre}')
    except FileExistsError:
        pass

    archivo = f'{common.config_path}{nombre}/horario.xlsx'
    wb.save(archivo)
    return archivo
    